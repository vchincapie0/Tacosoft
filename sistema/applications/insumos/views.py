from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from django.contrib import messages
from django.shortcuts import render
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
#Importacion de modelos y formularios

from .models import (
    ImplementosTrabajo,
    InsumosGenerico,
    ImplementosAudit,
)
from .forms import (
    ImplementosTrabajoForm,
    ImplementosUpdateForm,
    InsumosGenericoForm,
    InsumosGenericoUpdateForm,
    ImplementosAuditFilterForm,
    InsumosGenericoFilterForm,
)

from .models import ImplementosTrabajo,InsumosGenerico,ImplementosAudit
from .forms import  ImplementosUpdateForm,InsumosGenericoForm,InsumosGenericoUpdateForm,ImplementosAuditFilterForm


# Create your views here.

class InsumosGenericoListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de los Insumos Genericos'''
    model = InsumosGenerico
    template_name = "administrador/genericas/insumos/list_insumos_generico.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=10
    context_object_name = 'insumos'

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Obtener los parámetros de filtrado del formulario
        form = InsumosGenericoFilterForm(self.request.GET)

        # Aplicar filtros si el formulario es válido
        if form.is_valid():
            it_nombre = form.cleaned_data.get('it_nombre')

            # Filtrar por nombre de insumo genérico
            if it_nombre:
                queryset = queryset.filter(it_nombre__icontains=it_nombre)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = InsumosGenericoFilterForm(self.request.GET)
        return context
    
class InsumosGenericoCreateView(LoginRequiredMixin, CreateView):
    '''Clase donde se crea una nueva materia prima'''
    model = InsumosGenerico
    template_name = "administrador/genericas/insumos/add_insumos_generico.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class=InsumosGenericoForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('insumos_app:list_insumos_generico')

class InsumosGenericoUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para actualizar los datos de user'''
    model = InsumosGenerico
    template_name = "administrador/genericas/insumos/update_insumos_generico.html"
    login_url=reverse_lazy('users_app:login')
    form_class=InsumosGenericoUpdateForm
    success_url= reverse_lazy('insumos_app:list_insumos')

class InsumosGenericoDeleteView(LoginRequiredMixin, DeleteView):
    '''Vista para borrar Implenentos de Trabajo'''
    model = InsumosGenerico
    template_name = "administrador/genericas/insumos/delete_insumos_generico.html"
    login_url=reverse_lazy('users_app:login')
    success_url= reverse_lazy('insumos_app:list_insumos_generico')

class InsumosListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de los Implementos de trabajo'''
    model = ImplementosTrabajo
    template_name = "insumos/list_insumos.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=10
    context_object_name = 'insumos'

    def get_queryset(self):
        '''Funcion que toma de la barra de busqueda la pablabra clave para filtrar'''
        palabra_clave= self.request.GET.get("kword",'')
        lista = ImplementosTrabajo.objects.filter(
           it_nombre__it_nombre__icontains = palabra_clave
        )
        return lista
    
class ImplementosUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para actualizar los datos de user'''
    model = ImplementosTrabajo
    template_name = "insumos/update_insumos.html"
    login_url=reverse_lazy('users_app:login')
    form_class=ImplementosUpdateForm
    success_url= reverse_lazy('insumos_app:list_insumos')

    def form_valid(self, form):
        #Obtener los datos del fomulario
        nombre = form.cleaned_data['it_nombre']

        # Agregar un mensaje de éxito con el nombre de usuario
        messages.success(self.request, f'¡El implemento de trabajo {nombre} se ha actualizado correctamente!')

        return super(ImplementosUpdateView, self).form_valid(form)
    
class ImplementosDeleteView(LoginRequiredMixin, DeleteView):
    '''Vista para borrar Implenentos de Trabajo'''
    model = ImplementosTrabajo
    template_name = "insumos/delete_insumos.html"
    login_url=reverse_lazy('users_app:login')
    success_url= reverse_lazy('insumos_app:list_insumos')

class ImplementosAuditListView(LoginRequiredMixin, ListView):
    model= ImplementosAudit
    template_name='administrador/auditorias/implementosaudit.html'
    paginate_by=10
    context_object_name='auditoria'

    def get_queryset(self):
        queryset = super().get_queryset()

        # Obtener los parámetros de filtrado del formulario
        form = ImplementosAuditFilterForm(self.request.GET)

        # Aplicar filtros si el formulario es válido
        if form.is_valid():
            implementos = form.cleaned_data.get('implementos')
            action = form.cleaned_data.get('action')
            changed_by = form.cleaned_data.get('changed_by')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')

            # Filtrar por usuario, acción, usuario que realizó el cambio y rango de fechas
            if implementos:
                queryset = queryset.filter(implementos=implementos)
            if action:
                queryset = queryset.filter(action=action)
            if changed_by:
                queryset = queryset.filter(changed_by=changed_by)
            if start_date:
                queryset = queryset.filter(changed_at__gte=start_date)
            if end_date:
                # Agregar 1 día a la fecha final para incluir todos los registros de ese día
                end_date += timezone.timedelta(days=1)
                queryset = queryset.filter(changed_at__lt=end_date)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ImplementosAuditFilterForm(self.request.GET)
        return context

def export_implementos_to_excel(request):
    '''Vista para exportar datos de tabla insumos en formato excel'''
    # Obtener la fecha y hora actual
    fecha_descarga = timezone.now().strftime("%Y-%m-%d %H:%M:%S")

    # Obtener los datos de implementos que quieres exportar
    implementos = ImplementosTrabajo.objects.filter(deleted=False)  # Filtrar proveedores activos

    # Crear un nuevo libro de Excel y una hoja de trabajo
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Implementos'

    # Establecer estilos para la primera línea (encabezado personalizado)
    title_font = Font(bold=True)
    title_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # Transparente
    title_alignment = Alignment(horizontal='center')
    
    # Agregar fila de título personalizado
    worksheet.append(['TACO MAS'])  # Agregar texto del título
    worksheet.merge_cells('A1:D1')  # Combinar celdas para el título
    title_cell = worksheet['A1']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment

    # Agregar información adicional (fecha y nombre del software) en una nueva fila
    worksheet.append(['Fecha de descarga:', fecha_descarga])
    worksheet.append(['Software:', 'Tacosoft'])

    # Agregar espacio en blanco entre la información adicional y los encabezados
    worksheet.append([])  # Agregar una fila vacía

    # Agregar encabezados a la siguiente fila
    headers = ['Implemento', 'Cantidad', 'Fecha de Entrega','Estado']
    worksheet.append(headers)

    # Aplicar estilos a la fila de encabezados (fila actual + 2)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # Gris claro
    header_alignment = Alignment(horizontal='center')

    for col in range(1, len(headers) + 1):
        header_cell = worksheet.cell(row=worksheet.max_row, column=col)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment

    # Agregar datos de implementos a las siguientes filas
    for implemento in implementos:
        data_row = [
            implemento.it_nombre.it_nombre,  # Access related field's name
            implemento.it_cantidad,
            implemento.it_fechaEntrega.strftime("%Y-%m-%d"),  # Format date
            implemento.get_it_estado_display()  # Get display value of choice field
        ]
        worksheet.append(data_row)

    # Crear una respuesta HTTP con el archivo Excel como contenido
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=implementos.xlsx'

    # Guardar el libro de Excel en la respuesta HTTP
    workbook.save(response)

    return response

def export_implementos_to_csv(request):
    '''Vista para exportar datos de tabla implementos en formato CSV'''
    implementos = ImplementosTrabajo.objects.filter(deleted=False)  # Obtener datos de proveedores
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=implementos.csv'

    writer = csv.writer(response)
    writer.writerow(['Implemento', 'Cantidad', 'Fecha de Entrega','Estado'])  # Encabezados de columnas

    for implemento in implementos:
        writer.writerow([
            implemento.it_nombre.it_nombre,  # Access related field's name
            implemento.it_cantidad,
            implemento.it_fechaEntrega.strftime("%Y-%m-%d"),  # Format date
            implemento.get_it_estado_display()  # Get display value of choice field
        ])


    return response