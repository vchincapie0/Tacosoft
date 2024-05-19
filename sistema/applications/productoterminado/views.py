from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from django.contrib import messages
from django.shortcuts import render
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
#Importacion de modelos y formularios
from .models import (
    ProductoTerminado,
    ExistenciaPT,
    CaracteristicasOrganolepticasPT,
    EmpaqueProductoTerminado,
    Vacio,
    ProductoTerminadoGenerico,
    ProductoTerminadoGenerico,
    ProductoTerminadoAudit
    
)
from .forms import (
    ProductoTerminadoForm,
    CaracteristicasOrganolepticasPTForm,
    EmpaqueProductoTerminadoForm,
    VacioForm,
    CaracteristicasPTUpdateForm,
    EmpaqueUpdateForm,
    VacioUpdateForm,
    ProductoTerminadoGenericoForm,
    ProductoAuditFilterForm
)

# Create your views here.

class ProduListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de producto terminado'''
    model = ProductoTerminado
    template_name = "productoterminado/list_produ.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=10
    context_object_name = 'productoterminado'

    def get_queryset(self):
        '''Funcion que toma de la barra de busqueda la pablabra clave para filtrar'''
        palabra_clave= self.request.GET.get("kword",'')
        lista = ProductoTerminado.objects.filter(
            pt_nombre__pt_nombre__icontains = palabra_clave
        )
        return lista
    
class ProduCreateView(LoginRequiredMixin, CreateView):
    '''Clase donde se crea un producto terminado'''
    model = ProductoTerminado
    template_name = "productoterminado/add_produ.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = ProductoTerminadoForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('produ_app:list_produ')  

class ProduUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para actualizar los datos de producto terminado'''
    model = ProductoTerminado
    template_name = "productoterminado/update_produ.html"
    login_url=reverse_lazy('users_app:login')
    form_class=ProductoTerminadoForm
    success_url= reverse_lazy('produ_app:list_produ')
    
class ProduDeleteView(LoginRequiredMixin, DeleteView):
    '''Vista para borrar los producto terminado'''
    model = ProductoTerminado
    template_name = "productoterminado/delete_produ.html"
    login_url=reverse_lazy('users_app:login')
    success_url= reverse_lazy('produ_app:list_produ')

class ExistenciaPTView(LoginRequiredMixin, ListView):
    '''Vists para la creacion de la existencias producto terminado'''
    model = ExistenciaPT
    template_name = "productoterminado/existenciaPT.html"
    login_url=reverse_lazy('users_app:login')
    success_url= reverse_lazy('produ_app:exitenciaPT')

class CaracteristicasProductoTerminadoCreateView(LoginRequiredMixin, CreateView):
    '''Vista para la creacion de las caracteristicas organolepticas de Producto terminado'''
    model = CaracteristicasOrganolepticasPT
    template_name = "productoterminado/caracteristicas_PT.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = CaracteristicasOrganolepticasPTForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('produ_app:list_produ')

class CaracteristicasProductoTerminadoUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para la edición de las caracteristicas organolepticas de producto terminado'''
    model = CaracteristicasOrganolepticasPT
    template_name = "productoterminado/updatecaracteristicas_pt.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = CaracteristicasPTUpdateForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('produ_app:list_produ')

class ProductoTerminadoDetailView(LoginRequiredMixin, DetailView):
    
    '''Vista donde se muestran los detalles de producto terminado'''
    model = ProductoTerminado
    template_name = "productoterminado/detail_PT.html"
    login_url=reverse_lazy('users_app:login')
    context_object_name = 'productoterminado'

class EmpaqueProductoTerminadoCreateView(LoginRequiredMixin, CreateView):
    '''Vists para la creacion del empaque producto terminado'''
    model = EmpaqueProductoTerminado
    template_name = "productoterminado/empaque_pt.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = EmpaqueProductoTerminadoForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('produ_app:list_produ')
    
    def form_valid(self, form):
        '''funcion para automatizar el campo '''
        user = self.request.user
             # Guarda el formulario sin commit para asignar manualmente el usuario
        empaque = form.save(commit=False)
             # Asigna el usuario al campo pedi_user
        empaque.responsable = user
             # Ahora sí, guarda el pedido en la base de datos
        empaque.save()
        return super().form_valid(form)

class EmpaqueProductoTerminadoUpdateView(LoginRequiredMixin, UpdateView):
    '''Vists para la edición del empaque producto terminado'''
    model = EmpaqueProductoTerminado
    template_name = "productoterminado/empaqueupdate_pt.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = EmpaqueUpdateForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('produ_app:list_produ')
    
    def form_valid(self, form):
        '''funcion para automatizar el campo '''
        user = self.request.user
             # Guarda el formulario sin commit para asignar manualmente el usuario
        empaque = form.save(commit=False)
             # Asigna el usuario al campo pedi_user
        empaque.responsable = user
             # Ahora sí, guarda el pedido en la base de datos
        empaque.save()
        return super().form_valid(form)    
    
class VacioProductoTerminadoCreateView(LoginRequiredMixin, CreateView):
    '''Vists para la creacion del vacio producto terminado'''
    model = Vacio
    template_name = "productoterminado/vacio_pt.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = VacioForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('produ_app:list_produ')
    
    def form_valid(self, form):
        '''funcion para automatizar el campo '''
        user = self.request.user
             # Guarda el formulario sin commit para asignar manualmente el usuario
        Vacio = form.save(commit=False)
             # Asigna el usuario al campo pedi_user
        Vacio.responsable = user
             # Ahora sí, guarda el pedido en la base de datos
        Vacio.save()
        return super().form_valid(form)

class VacioProductoTerminadoUpdateView(LoginRequiredMixin, UpdateView):
    '''Vists para la edición del vacio producto terminado'''
    model = Vacio
    template_name = "productoterminado/vacioupdate_pt.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = VacioUpdateForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('produ_app:list_produ')
    
    def form_valid(self, form):
        '''funcion para automatizar el campo '''
        user = self.request.user
             # Guarda el formulario sin commit para asignar manualmente el usuario
        Vacio = form.save(commit=False)
             # Asigna el usuario al campo pedi_user
        Vacio.responsable = user
             # Ahora sí, guarda el pedido en la base de datos
        Vacio.save()
        return super().form_valid(form)   

class ProductoTerminadoGenericoListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de las materias primas'''
    model = ProductoTerminadoGenerico
    template_name = "administrador/genericas/list_pt_generico.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=10
    context_object_name = 'producto'

    def get_queryset(self):
        palabra_clave = self.request.GET.get("kword", '')
        
        # Filtrar por nombre específico de la materia prima
        queryset = ProductoTerminadoGenerico.objects.filter(
            pt_nombre__icontains=palabra_clave
        )
        
        return queryset   
    
class ProductoTerminadoGenericoCreateView(LoginRequiredMixin, CreateView):
    '''Clase donde se crea un nuevo Producto terminado'''
    model = ProductoTerminadoGenerico
    template_name = "administrador/genericas/add_pt_generico.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = ProductoTerminadoGenericoForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('produ_app:list_pt_generico')

class ProductoTerminadoGenericoUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para actualizar los datos  Producto terminado'''
    model = ProductoTerminadoGenerico
    template_name = "administrador/genericas/update_pt_generico.html"
    login_url=reverse_lazy('users_app:login')
    form_class=ProductoTerminadoGenericoForm
    success_url= reverse_lazy('produ_app:list_pt_generico')

class ProductoTerminadoGenericoDeleteView(LoginRequiredMixin, DeleteView):
    '''Vista para borrar Producto terminado'''
    model = ProductoTerminadoGenerico
    template_name = "administrador/genericas/delete_pt_generico.html"
    login_url=reverse_lazy('users_app:login')
    success_url= reverse_lazy('produ_app:list_pt_generico')

class ProductoAuditListView(LoginRequiredMixin, ListView):
    model= ProductoTerminadoAudit
    template_name='administrador/auditorias/productoaudit.html'
    paginate_by=10
    context_object_name='auditoria'

    def get_queryset(self):
        queryset = super().get_queryset()

        # Obtener los parámetros de filtrado del formulario
        form = ProductoAuditFilterForm(self.request.GET)

        # Aplicar filtros si el formulario es válido
        if form.is_valid():
            productoterminado = form.cleaned_data.get('productoterminado')
            action = form.cleaned_data.get('action')
            changed_by = form.cleaned_data.get('changed_by')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')

            # Filtrar por usuario, acción, usuario que realizó el cambio y rango de fechas
            if productoterminado:
                queryset = queryset.filter(productoterminado=productoterminado)
            if action:
                queryset = queryset.filter(action=action)
            if changed_by:
                queryset = queryset.filter(changed_by=changed_by)
            if start_date:
                queryset = queryset.filter(changed_at__gte=start_date)
            if end_date:
                # Agregar 1 día a la fecha final para incluir todos los registros de ese día
                end_date += timezone.timedelta(days=1)
                queryset = queryset.filter(changed_at__lt=end_date)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ProductoAuditFilterForm(self.request.GET)
        return context
    
def export_productoterminado_to_excel(request):
    '''Vista para exportar datos de tabla producto terminado en formato excel'''
    # Obtener la fecha y hora actual
    fecha_descarga = timezone.now().strftime("%Y-%m-%d %H:%M:%S")

    # Obtener los datos de producto terminado que quieres exportar
    productoterminado = ProductoTerminado.objects.filter(deleted=False)  # Filtrar productos terminados activos

    # Crear un nuevo libro de Excel y una hoja de trabajo
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Producto Terminado'

    # Establecer estilos para la primera línea (encabezado personalizado)
    title_font = Font(bold=True)
    title_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # Transparente
    title_alignment = Alignment(horizontal='center')
    
    # Agregar fila de título personalizado
    worksheet.append(['TACO MAS'])  # Agregar texto del título
    worksheet.merge_cells('A1:C1')  # Combinar celdas para el título
    title_cell = worksheet['A1']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment

    # Agregar información adicional (fecha y nombre del software) en una nueva fila
    worksheet.append(['Fecha de descarga:', fecha_descarga])
    worksheet.append(['Software:', 'Tacosoft'])

    # Agregar espacio en blanco entre la información adicional y los encabezados
    worksheet.append([])  # Agregar una fila vacía

    # Agregar encabezados a la siguiente fila
    headers = ['Lote', 'Producto', 'Cantidad','Fecha de preparación','Fecha de vencimiento']
    worksheet.append(headers)

    # Aplicar estilos a la fila de encabezados (fila actual + 2)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # Gris claro
    header_alignment = Alignment(horizontal='center')

    for col in range(1, len(headers) + 1):
        header_cell = worksheet.cell(row=worksheet.max_row, column=col)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment

    # Agregar datos de proveedores a las siguientes filas
    for productoterminado in ProductoTerminado:
        data_row = [productoterminado.pt_lote, productoterminado.pt_nombre, productoterminado.pt_fechapreparacion,productoterminado.pt_fechavencimiento]
        worksheet.append(data_row)

    # Crear una respuesta HTTP con el archivo Excel como contenido
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productoterminado.xlsx'

    # Guardar el libro de Excel en la respuesta HTTP
    workbook.save(response)

    return response

def export_productoterminado_to_csv(request):
    '''Vista para exportar datos de tabla producto terminado en formato CSV'''
    productoterminado = ProductoTerminado.objects.filter(deleted=False)  # Obtener datos de proveedores
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=productoterminado.csv'

    writer = csv.writer(response)
    writer.writerow(['Lote', 'Producto', 'Cantidad','Fecha de preparación','Fecha de vencimiento'])  # Encabezados de columnas

    for productoterminado in ProductoTerminado:
        writer.writerow([productoterminado.pt_lote, productoterminado.pt_nombre, productoterminado.pt_fechapreparacion,productoterminado.pt_fechavencimiento])

    return response